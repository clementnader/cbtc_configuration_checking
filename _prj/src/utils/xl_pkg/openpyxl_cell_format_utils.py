#!/usr/bin/env python
# -*- coding: utf-8 -*-

import numpy
import openpyxl.styles as xl_styles
import openpyxl.styles.borders as xl_borders
from .xl_utils import *
from ..common_utils import *


__all__ = ["xl_borders", "XlAlign",
           "create_cell", "create_merged_cell",
           "set_fixed_number_of_digits",
           "set_font_size", "set_bold_font", "set_bg_color",
           "set_vertical_alignment", "set_horizontal_alignment", "enable_line_wrap",
           "adjust_fixed_row_height",
           "draw_exterior_borders", "draw_all_borders_of_a_range",
           "draw_exterior_borders_of_a_range"]


class XlAlign:
    center = "center"
    # Horizontal Alignments
    general = "general"
    left = "left"
    right = "right"
    # Vertical Alignments
    top = "top"
    bottom = "bottom"


# ------ Create cells ------ #

def create_cell(ws: xl_ws.Worksheet, value: Union[str, float, None],
                cell: str = None, row: int = None, column: Union[str, int] = None,
                font_size: int = None, bold: bool = False, italic: bool = False, bg_color: str = None,
                align_vertical: str = XlAlign.center, align_horizontal: str = XlAlign.general, line_wrap: bool = False,
                nb_of_digits: int = None, borders: bool = False) -> None:
    row, column = get_row_and_column_from_cell(cell, row, column)

    ws.cell(row=row, column=column, value=value)
    # Cell Formatting
    _update_cell_formatting(ws, row, column, font_size, bold, italic, bg_color,
                            align_vertical, align_horizontal, line_wrap, nb_of_digits)
    # Cell Borders
    if borders:
        draw_exterior_borders(ws, row=row, column=column)


def create_merged_cell(ws: xl_ws.Worksheet, value: Union[str, float, None],
                       start_cell: str = None, start_row: int = None, start_column: Union[str, int] = None,
                       end_cell: str = None, end_row: int = None, end_column: Union[str, int] = None,
                       cell_range: str = None,
                       font_size: int = None, bold: bool = False, italic: bool = False, bg_color: str = None,
                       align_vertical: str = XlAlign.center, align_horizontal: str = XlAlign.general,
                       line_wrap: bool = True, nb_of_digits: int = None,
                       borders: bool = False, border_style: str = xl_borders.BORDER_THIN) -> None:
    if cell_range is not None:
        start_cell, end_cell = cell_range.split(":")
    start_row, start_column = get_row_and_column_from_cell(start_cell, start_row, start_column)
    end_row, end_column = get_row_and_column_from_cell(end_cell, end_row, end_column)

    ws.merge_cells(start_row=start_row, end_row=end_row,
                   start_column=start_column, end_column=end_column)
    ws.cell(row=start_row, column=start_column, value=value)
    # Cell Formatting
    _update_cell_formatting(ws, start_row, start_column, font_size, bold, italic, bg_color,
                            align_vertical, align_horizontal, line_wrap, nb_of_digits)
    # Cell Borders
    if borders:
        draw_all_borders_of_a_range(ws, start_row=start_row, end_row=end_row, start_column=start_column,
                                    end_column=end_column, border_style=border_style)


def _update_cell_formatting(ws: xl_ws.Worksheet, row: int, column: int,
                            font_size: int, bold: bool, italic: bool, bg_color: str,
                            align_vertical: str, align_horizontal: str,
                            line_wrap: bool, nb_of_digits: int) -> None:
    if font_size is not None:
        set_font_size(ws, font_size=font_size, row=row, column=column)
    if bold:
        set_bold_font(ws, row=row, column=column)
    if italic:
        set_italic_font(ws, row=row, column=column)
    if bg_color is not None:
        set_bg_color(ws, hex_color=bg_color, row=row, column=column)
    if align_vertical:
        set_vertical_alignment(ws, align_vertical=align_vertical, row=row, column=column)
    if align_horizontal:
        set_horizontal_alignment(ws, align_horizontal=align_horizontal, row=row, column=column)
    if line_wrap:
        enable_line_wrap(ws, row=row, column=column)
    if nb_of_digits:
        set_fixed_number_of_digits(ws, number_of_digits=nb_of_digits, row=row, column=column)


# ------ Cell Formatting ------ #

def set_font_size(ws: xl_ws.Worksheet, font_size: int,
                  cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    font_style = xl_styles.Font(size=font_size)
    ws[cell].font = font_style


def set_bold_font(ws: xl_ws.Worksheet,
                  cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    font_style = xl_styles.Font(bold=True)
    ws[cell].font += font_style


def set_italic_font(ws: xl_ws.Worksheet,
                    cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    font_style = xl_styles.Font(italic=True)
    ws[cell].font += font_style


def set_bg_color(ws: xl_ws.Worksheet, hex_color: str,
                 cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    pattern_style = xl_styles.PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    ws[cell].fill = pattern_style


def set_vertical_alignment(ws: xl_ws.Worksheet, align_vertical: str = XlAlign.center,
                           cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    alignment_style = xl_styles.Alignment(vertical=align_vertical)
    ws[cell].alignment += alignment_style


def set_horizontal_alignment(ws: xl_ws.Worksheet, align_horizontal: str = XlAlign.general,
                             cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    alignment_style = xl_styles.Alignment(horizontal=align_horizontal)
    ws[cell].alignment += alignment_style


def enable_line_wrap(ws: xl_ws.Worksheet,
                     cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    alignment_style = xl_styles.Alignment(wrap_text=True)
    ws[cell].alignment += alignment_style


def adjust_fixed_row_height(ws: xl_ws.Worksheet,
                            cell: str = None, row: int = None, column: Union[str, int] = None,
                            line_length: int = None) -> None:
    # Used in order to work with formula inside cell, with which the text_wrap will not work automatically
    row, column = get_row_and_column_from_cell(cell, row, column)
    value = get_xlsx_value(ws, row, column)
    if value is None:
        return
    line_feed_count = value.count("\n")
    if line_length is None:
        automatic_wrap = 2
    else:
        column_width = ws.column_dimensions[get_xl_column_letter(column)].width
        automatic_wrap = numpy.ceil(line_length/column_width*4/5) - 1
    ws.row_dimensions[row].height = 15 * (line_feed_count + automatic_wrap + 1)


# ------ Number Format ------ #

def set_fixed_number_of_digits(ws: xl_ws.Worksheet, number_of_digits: int,
                               cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    number_format = "0." + "0"*number_of_digits
    ws[cell].number_format = number_format


def set_text_format(ws, row, column):
    cell = get_cell_from_row_and_column(row, column)
    ws[cell].number_format = "@"  # force text formatting


# ------ Cell Borders ------ #

def draw_exterior_borders(ws: xl_ws.Worksheet,
                          cell: str = None, row: int = None, column: Union[str, int] = None,
                          border_style: str = xl_borders.BORDER_THIN) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    side = xl_borders.Side(border_style=border_style)
    border = xl_borders.Border(left=side, right=side, top=side, bottom=side)
    ws[cell].border = border


def draw_all_borders_of_a_range(ws: xl_ws.Worksheet, cell_range: str = None,
                                start_row: int = None, end_row: int = None,
                                start_column: Union[int, str] = None, end_column: Union[int, str] = None,
                                border_style: str = xl_borders.BORDER_THIN) -> None:
    cell_range = get_cell_range(cell_range, start_row, end_row, start_column, end_column)
    side = xl_borders.Side(border_style=border_style)
    border = xl_borders.Border(left=side, right=side, top=side, bottom=side)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


def draw_exterior_borders_of_a_range(ws: xl_ws.Worksheet, cell_range: str = None,
                                     start_row: int = None, end_row: int = None,
                                     start_column: Union[int, str] = None, end_column: Union[int, str] = None,
                                     exterior_border_style: str = xl_borders.BORDER_MEDIUM,
                                     interior_border_style: str = xl_borders.BORDER_THIN) -> None:
    cell_range = get_cell_range(cell_range, start_row, end_row, start_column, end_column)
    # Remove all borders
    no_border = xl_borders.Border()
    for row in ws[cell_range]:
        for cell in row:
            cell.border = no_border
    # Add exterior borders
    ext_side = xl_borders.Side(border_style=exterior_border_style)
    left_border = xl_borders.Border(left=ext_side)
    right_border = xl_borders.Border(right=ext_side)
    top_border = xl_borders.Border(top=ext_side)
    bottom_border = xl_borders.Border(bottom=ext_side)
    # Add top border on first row
    first_row = ws[cell_range][0]
    for cell in first_row:
        cell.border += top_border
    # Add bottom border on last row
    last_row = ws[cell_range][-1]
    for cell in last_row:
        cell.border += bottom_border
    # Add left border on first column, and right border on last column
    for row in ws[cell_range]:
        first_cell = row[0]
        first_cell.border += left_border
        last_cell = row[-1]
        last_cell.border += right_border
    # Interior borders
    int_side = xl_borders.Side(border_style=interior_border_style)
    interior_border = xl_borders.Border(left=int_side, right=int_side, top=int_side, bottom=int_side)
    for row in ws[cell_range]:
        for cell in row:
            cell.border += interior_border  # add interior borders everywhere else
