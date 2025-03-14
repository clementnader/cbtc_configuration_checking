#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import warnings
import openpyxl
import openpyxl.utils as xl_ut
import openpyxl.worksheet.worksheet as xl_ws
import xlrd
from ..common_utils import *
from ..colors_pkg import *


__all__ = ["openpyxl", "xlrd", "xl_ut", "xl_ws", "load_xlsx_wb", "load_xlrd_wb", "get_xlrd_column", "get_xlrd_row",
           "get_xl_column_letter", "get_xl_column_number", "get_xlrd_float_value", "get_xlrd_value", "get_xlsx_value",
           "get_row_and_column_from_cell", "get_cell_from_row_and_column",
           "get_xl_max_row", "get_xl_max_column", "get_cell_range",
           "get_xl_col_from_number_or_letter"]


def load_xlsx_wb(xl_file_address: str, template: bool = False, read_only: bool = False) -> openpyxl.workbook.Workbook:
    warnings.filterwarnings("ignore", message="Cannot parse header or footer so it will be ignored",
                            category=UserWarning, module="openpyxl")  # deactivate the warning for header/footer
    warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed",
                            category=UserWarning, module="openpyxl")  # deactivate the warning for extension
    if template:
        # It seems that rich_text argument is causing some issues on the saving depending on openpyxl version.
        # wb = openpyxl.load_workbook(xl_file_address, rich_text=True)
        wb = openpyxl.load_workbook(xl_file_address)
    elif read_only:
        wb = openpyxl.load_workbook(xl_file_address, data_only=True, read_only=True, keep_links=False)
    else:
        # read-only active prevents the max_row to work
        wb = openpyxl.load_workbook(xl_file_address, data_only=True, read_only=False, keep_links=False)
    warnings.filterwarnings("default")
    return wb


def load_xlrd_wb(xl_file_address: str, formatting_info=False, on_demand=False) -> openpyxl.workbook.Workbook:
    # formatting_info is useful to know if a cell is in percentage mode
    try:
        wb = xlrd.open_workbook(xl_file_address, formatting_info=formatting_info, on_demand=on_demand)
    except AssertionError:
        print_warning(f"Error reading the Excel file {Color.yellow}\"{xl_file_address}\"{Color.reset} "
                      f"with xlrd library.\n"
                      f"We use Excel Application to re-save it as another file to try to fix the issue.")
        new_xl_file_address = "_patch".join(os.path.splitext(xl_file_address))

        import comtypes.client
        prog_id = "Excel.Application"
        xl = comtypes.client.CreateObject(prog_id)
        wb = xl.Workbooks.Open(xl_file_address)
        try:
            wb.SaveAs(new_xl_file_address, FileFormat=56)  # Excel 97-2003 Workbook (*.xls)
            # see https://learn.microsoft.com/en-us/office/vba/api/excel.xlfileformat
            xl.DisplayAlerts = False
            xl.Quit()
            print_log(f"File has been re-saved as {Color.default}\"{new_xl_file_address}\"{Color.reset}.\n"
                      f"We try to read this new file with xlrd library.")

        except comtypes.COMError:  # We can have a COM error if the file is already created and the user chooses
            # to not replace it. The execution can continue with the existing file.
            print_log(f"A file {Color.default}\"{new_xl_file_address}\"{Color.reset} already exists.\n"
                      f"We take this file and try to read it with xlrd library.")

        wb = xlrd.open_workbook(new_xl_file_address, formatting_info=formatting_info, on_demand=on_demand)
        print_log(f"It has worked. Execution continues.")

    return wb


def get_xlrd_column(column_number: int) -> int:
    return column_number - 1


def get_xlrd_row(row: int) -> int:
    return row - 1


def get_xl_column_letter(column_number: int) -> str:
    return xl_ut.get_column_letter(column_number)


def get_xl_column_number(column_letter: str) -> int:
    return xl_ut.column_index_from_string(column_letter)


def get_xlrd_float_value(ws: xlrd.sheet, row: int, column: int) -> Optional[Union[float, str]]:
    value = get_xlrd_value(ws, row, column)
    if isinstance(value, str):
        try:
            value = float(value.replace(",", "."))
        except ValueError:
            pass
    return value


def get_xlrd_value(ws: xlrd.sheet, row: int, column: int) -> Optional[str]:
    xlrd_row = get_xlrd_row(row)
    xlrd_col = get_xlrd_column(column)
    try:
        cell_value = ws.cell_value(xlrd_row, xlrd_col)
    except IndexError:
        cell_value = None
    if cell_value == "":
        cell_value = None
    return cell_value


def get_xlsx_value(ws, row: int, column: int) -> Optional[str]:
    cell_value = ws.cell(row=row, column=column).value
    if cell_value == "":
        cell_value = None
    return cell_value


def get_row_and_column_from_cell(cell: str = None, row: int = None, column: Union[str, int] = None) -> tuple[int, int]:
    if cell is not None:
        row, column = split_cell(cell)
    if isinstance(column, str):
        column = xl_ut.column_index_from_string(column)
    return row, column


def split_cell(cell: str) -> tuple[int, str]:
    first_digit_index = -1
    for i, character in enumerate(cell):
        if character.isdigit():
            first_digit_index = i
            break
    column = cell[:first_digit_index]
    row = int(cell[first_digit_index:])
    return row, column


def get_cell_from_row_and_column(cell: str = None, row: int = None, column: Union[str, int] = None) -> str:
    if cell is None:
        if isinstance(column, int):
            column = xl_ut.get_column_letter(column)
        cell = f"{column}{row}"
    return cell


def get_xl_max_row() -> int:
    max_row = openpyxl.xml.constants.MAX_ROW  # MAX_ROW = 1048576 = 2^20
    return max_row


def get_xl_max_column() -> int:
    max_column = openpyxl.xml.constants.MAX_COLUMN  # MAX_COLUMN = 16384 = 2^14
    return max_column


def get_cell_range(cell_range: str = None, start_row: int = None, end_row: int = None,
                   start_column: Union[int, str] = None, end_column: Union[int, str] = None) -> str:
    if cell_range is None:
        if start_row is None:
            start_row = 1
        if start_column is None:
            start_column = 1
        start_cell = get_cell_from_row_and_column(row=start_row, column=start_column)

        if end_row is None:
            end_row = get_xl_max_row()
        if end_column is None:
            end_column = get_xl_max_column()
        end_cell = get_cell_from_row_and_column(row=end_row, column=end_column)

        cell_range = f"{start_cell}:{end_cell}"
    return cell_range


def get_xl_col_from_number_or_letter(val: str) -> int:
    try:
        val = int(val)
    except ValueError:
        val = get_xl_column_number(val)
    return val
