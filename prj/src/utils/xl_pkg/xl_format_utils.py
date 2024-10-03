#!/usr/bin/env python
# -*- coding: utf-8 -*-

import numpy
import openpyxl.comments as xl_comments
import openpyxl.drawing.image as xl_image
import openpyxl.styles as xl_styles
import openpyxl.styles.differential as xl_diff_styles
import openpyxl.styles.borders as xl_borders
import openpyxl.formatting as xl_format
import openpyxl.formatting.rule as xl_format_rule
import openpyxl.worksheet.worksheet
import openpyxl.workbook.defined_name as xl_defined_name
from .xl_utils import *
from ..common_utils import *


__all__ = ["xl_borders", "XlAlign", "XlFontColor", "XlBgColor", "get_xl_bg_dimmer_color", "add_cell_comment",
           "add_image",
           "create_cell", "create_merged_cell",
           "set_fixed_number_of_digits",
           "set_font_size", "set_bold_font", "set_bg_color",
           "set_vertical_alignment", "set_horizontal_alignment", "enable_line_wrap",
           "adjust_fixed_row_height",
           "draw_exterior_borders", "draw_all_borders_of_a_range",
           "draw_exterior_borders_of_a_range",
           "add_unique_values_conditional_formatting", "add_duplicate_values_conditional_formatting",
           "add_is_equal_conditional_formatting", "add_formula_conditional_formatting",
           "create_defined_name"]


class XlAlign:
    center = "center"
    # Horizontal Alignments
    general = "general"
    left = "left"
    right = "right"
    # Vertical Alignments
    top = "top"
    bottom = "bottom"


class XlFontColor:
    # status colors
    ko = "9C0006"
    ok = "006100"
    na = "9C5700"
    # other colors
    dark_red = "9C0006"
    orange = "E97132"


class XlBgColor:
    # status colors
    ko = "FFC7CE"
    ok = "C6EFCE"
    na = "FFEB9C"
    # main colors
    yellow = "FFD050"
    light_yellow = "FFFFBB"
    green = "A0E0A0"
    light_green = "DDFFCC"
    blue = "9BC2E6"
    light_blue = "DDEBF7"
    # other colors
    light_orange = "FFCC99"
    light_pink = "FFCCCC"
    light_red = "FF9999"
    light_blue2 = "CCFFFF"
    light_blue3 = "99CCFF"
    special_blue = "9999FF"
    # grey
    grey = "BFBFBF"
    light_grey = "D9D9D9"


def get_xl_bg_dimmer_color(bg: str):
    xl_bg_colors_dict = get_class_attr_dict(XlBgColor)
    for key, val in xl_bg_colors_dict.items():
        if bg == val:
            if key.startswith("light_"):
                dimmer_key = key.removeprefix("light_")
                if dimmer_key in xl_bg_colors_dict:
                    return xl_bg_colors_dict[dimmer_key]
    return bg


# ------ Cell Comment ------ #

def add_cell_comment(ws: openpyxl.worksheet.worksheet.Worksheet, comment: str,
                     cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    comment = xl_comments.Comment(comment, None)
    ws[cell].comment = comment


# ------ Image ------ #

def add_image(ws: openpyxl.worksheet.worksheet.Worksheet, image_path: str,
              cell: str = None, row: int = None, column: Union[str, int] = None,
              ratio: float = 1.) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    image = xl_image.Image(image_path)
    image.height *= ratio
    image.width *= ratio
    ws.add_image(image, anchor=cell)


# ------ Create cells ------ #

def create_cell(ws: openpyxl.worksheet.worksheet.Worksheet, value: Union[str, float, None],
                cell: str = None, row: int = None, column: Union[str, int] = None,
                font_size: int = None, bold: bool = False, italic: bool = False, bg_color: str = None,
                align_vertical: str = XlAlign.center, align_horizontal: str = XlAlign.general, line_wrap: bool = False,
                nb_of_digits: int = None, borders: bool = False) -> None:
    row, column = get_row_and_column_from_cell(cell, row, column)

    ws.cell(row=row, column=column, value=value)
    # Cell Formatting
    _update_cell_formatting(ws, row, column, font_size, bold, italic, bg_color,
                            align_vertical, align_horizontal, line_wrap, nb_of_digits)
    # Cell Borders
    if borders:
        draw_exterior_borders(ws, row=row, column=column)


def create_merged_cell(ws: openpyxl.worksheet.worksheet.Worksheet, value: str,
                       start_cell: str = None, start_row: int = None, start_column: Union[str, int] = None,
                       end_cell: str = None, end_row: int = None, end_column: Union[str, int] = None,
                       font_size: int = None, bold: bool = False, italic: bool = False, bg_color: str = None,
                       align_vertical: str = XlAlign.center, align_horizontal: str = XlAlign.general,
                       line_wrap: bool = True, nb_of_digits: int = None,
                       borders: bool = False, border_style: str = xl_borders.BORDER_THIN) -> None:
    start_row, start_column = get_row_and_column_from_cell(start_cell, start_row, start_column)
    end_row, end_column = get_row_and_column_from_cell(end_cell, end_row, end_column)

    ws.merge_cells(start_row=start_row, end_row=end_row,
                   start_column=start_column, end_column=end_column)
    ws.cell(row=start_row, column=start_column, value=value)
    # Cell Formatting
    _update_cell_formatting(ws, start_row, start_column, font_size, bold, italic, bg_color,
                            align_vertical, align_horizontal, line_wrap, nb_of_digits)
    # Cell Borders
    if borders:
        draw_all_borders_of_a_range(ws, start_row=start_row, end_row=end_row, start_column=start_column,
                                    end_column=end_column, border_style=border_style)


def _update_cell_formatting(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, column: int,
                            font_size: int, bold: bool, italic: bool, bg_color: str,
                            align_vertical: str, align_horizontal: str,
                            line_wrap: bool, nb_of_digits: int) -> None:
    if font_size is not None:
        set_font_size(ws, font_size=font_size, row=row, column=column)
    if bold:
        set_bold_font(ws, row=row, column=column)
    if italic:
        set_italic_font(ws, row=row, column=column)
    if bg_color is not None:
        set_bg_color(ws, hex_color=bg_color, row=row, column=column)
    if align_vertical:
        set_vertical_alignment(ws, align_vertical=align_vertical, row=row, column=column)
    if align_horizontal:
        set_horizontal_alignment(ws, align_horizontal=align_horizontal, row=row, column=column)
    if line_wrap:
        enable_line_wrap(ws, row=row, column=column)
    if nb_of_digits:
        set_fixed_number_of_digits(ws, number_of_digits=nb_of_digits, row=row, column=column)


# ------ Number Formatting ------ #

def set_fixed_number_of_digits(ws: openpyxl.worksheet.worksheet.Worksheet, number_of_digits: int,
                               cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    number_format = "0." + "0"*number_of_digits
    ws[cell].number_format = number_format


# ------ Cell Formatting ------ #

def set_font_size(ws: openpyxl.worksheet.worksheet.Worksheet, font_size: int,
                  cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    font_style = openpyxl.styles.Font(size=font_size)
    ws[cell].font = font_style


def set_bold_font(ws: openpyxl.worksheet.worksheet.Worksheet,
                  cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    font_style = openpyxl.styles.Font(bold=True)
    ws[cell].font += font_style


def set_italic_font(ws: openpyxl.worksheet.worksheet.Worksheet,
                    cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    font_style = openpyxl.styles.Font(italic=True)
    ws[cell].font += font_style


def set_bg_color(ws: openpyxl.worksheet.worksheet.Worksheet, hex_color: str,
                 cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    pattern_style = xl_styles.PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    ws[cell].fill = pattern_style


def set_vertical_alignment(ws: openpyxl.worksheet.worksheet.Worksheet, align_vertical: str = XlAlign.center,
                           cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    alignment_style = xl_styles.Alignment(vertical=align_vertical)
    ws[cell].alignment += alignment_style


def set_horizontal_alignment(ws: openpyxl.worksheet.worksheet.Worksheet, align_horizontal: str = XlAlign.general,
                             cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    alignment_style = xl_styles.Alignment(horizontal=align_horizontal)
    ws[cell].alignment += alignment_style


def enable_line_wrap(ws: openpyxl.worksheet.worksheet.Worksheet,
                     cell: str = None, row: int = None, column: Union[str, int] = None) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    alignment_style = openpyxl.styles.Alignment(wrap_text=True)
    ws[cell].alignment += alignment_style


def adjust_fixed_row_height(ws: openpyxl.worksheet.worksheet.Worksheet,
                            cell: str = None, row: int = None, column: Union[str, int] = None,
                            line_length: int = None) -> None:
    # Used to work with formula inside cell, with which the text_wrap will not work automatically
    row, column = get_row_and_column_from_cell(cell, row, column)
    value = get_xlsx_value(ws, row, column)
    if value is None:
        return
    line_feed_count = value.count("\n")
    if line_length is None:
        automatic_wrap = 2
    else:
        column_width = ws.column_dimensions[get_xl_column_letter(column)].width
        automatic_wrap = numpy.ceil(line_length/column_width*4/5) - 1
    ws.row_dimensions[row].height = 15 * (line_feed_count + automatic_wrap + 1)


# ------ Cell Borders ------ #

def draw_exterior_borders(ws: openpyxl.worksheet.worksheet.Worksheet,
                          cell: str = None, row: int = None, column: Union[str, int] = None,
                          border_style: str = xl_borders.BORDER_THIN) -> None:
    cell = get_cell_from_row_and_column(cell, row, column)
    side = xl_borders.Side(border_style=border_style)
    border = xl_borders.Border(left=side, right=side, top=side, bottom=side)
    ws[cell].border = border


def draw_all_borders_of_a_range(ws: openpyxl.worksheet.worksheet.Worksheet, cell_range: str = None,
                                start_row: int = None, end_row: int = None,
                                start_column: Union[int, str] = None, end_column: Union[int, str] = None,
                                border_style: str = xl_borders.BORDER_THIN) -> None:
    cell_range = get_cell_range(cell_range, start_row, end_row, start_column, end_column)
    side = xl_borders.Side(border_style=border_style)
    border = xl_borders.Border(left=side, right=side, top=side, bottom=side)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


def draw_exterior_borders_of_a_range(ws: openpyxl.worksheet.worksheet.Worksheet, cell_range: str = None,
                                     start_row: int = None, end_row: int = None,
                                     start_column: Union[int, str] = None, end_column: Union[int, str] = None,
                                     exterior_border_style: str = xl_borders.BORDER_MEDIUM,
                                     interior_border_style: str = xl_borders.BORDER_THIN) -> None:
    cell_range = get_cell_range(cell_range, start_row, end_row, start_column, end_column)
    # Remove all borders
    no_border = xl_borders.Border()
    for row in ws[cell_range]:
        for cell in row:
            cell.border = no_border
    # Add exterior borders
    ext_side = xl_borders.Side(border_style=exterior_border_style)
    left_border = xl_borders.Border(left=ext_side)
    right_border = xl_borders.Border(right=ext_side)
    top_border = xl_borders.Border(top=ext_side)
    bottom_border = xl_borders.Border(bottom=ext_side)
    # Add top border on first row
    first_row = ws[cell_range][0]
    for cell in first_row:
        cell.border += top_border
    # Add bottom border on last row
    last_row = ws[cell_range][-1]
    for cell in last_row:
        cell.border += bottom_border
    # Add left border on first column, and right border on last column
    for row in ws[cell_range]:
        first_cell = row[0]
        first_cell.border += left_border
        last_cell = row[-1]
        last_cell.border += right_border
    # Interior borders
    int_side = xl_borders.Side(border_style=interior_border_style)
    interior_border = xl_borders.Border(left=int_side, right=int_side, top=int_side, bottom=int_side)
    for row in ws[cell_range]:
        for cell in row:
            cell.border += interior_border  # add interior borders everywhere else


# ------ Conditional Formatting ------ #

# xl_format.Rule
#   rule type: ['expression', 'cellIs', 'colorScale', 'dataBar', 'iconSet', 'top10', 'uniqueValues',
# 'duplicateValues', 'containsText', 'notContainsText', 'beginsWith', 'endsWith', 'containsBlanks',
# 'notContainsBlanks', 'containsErrors', 'notContainsErrors', 'timePeriod', 'aboveAverage']
#   rule priority: int
#   rule stopIfTrue: bool
#   rule operator: ['lessThan', 'lessThanOrEqual', 'equal', 'notEqual', 'greaterThanOrEqual', 'greaterThan',
# 'between', 'notBetween', 'containsText', 'notContains', 'beginsWith', 'endsWith']
#   rule formula: list[str]
#   rule dxf: DifferentialStyle

def add_unique_values_conditional_formatting(ws: openpyxl.worksheet.worksheet.Worksheet, cell_range: str,
                                             font_color: str, bg_color: str, bold: bool = False) -> None:
    dxf = _get_differential_style(font_color, bg_color, bold)
    if dxf is None:
        return
    rule = xl_format.Rule(type="uniqueValues", priority=1, dxf=dxf)
    ws.conditional_formatting.add(cell_range, rule)


def add_duplicate_values_conditional_formatting(ws: openpyxl.worksheet.worksheet.Worksheet, cell_range: str,
                                                font_color: str, bg_color: str, bold: bool = False) -> None:
    dxf = _get_differential_style(font_color, bg_color, bold)
    if dxf is None:
        return
    rule = xl_format.Rule(type="duplicateValues", priority=1, dxf=dxf)
    ws.conditional_formatting.add(cell_range, rule)


def add_is_equal_conditional_formatting(ws: openpyxl.worksheet.worksheet.Worksheet, cell_range: str,
                                        value: Union[str, float, int],
                                        font_color: str = None, bg_color: str = None, bold: bool = False) -> None:
    dxf = _get_differential_style(font_color, bg_color, bold)
    if dxf is None:
        return
    formula_rule = xl_format.Rule(type="cellIs", operator="equal", formula=[value], dxf=dxf)
    ws.conditional_formatting.add(cell_range, formula_rule)


def _get_font_n_fill(font_color: str = None, bg_color: str = None, bold: bool = False):
    if font_color is not None and bold:
        font = xl_styles.Font(color=font_color, bold=True)
    elif font_color is not None:
        font = xl_styles.Font(color=font_color)
    elif bold:
        font = xl_styles.Font(bold=True)
    else:
        font = None
    if bg_color is not None:
        fill = xl_styles.PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    else:
        fill = None
    return font, fill


def _get_differential_style(font_color: str = None, bg_color: str = None, bold: bool = False):
    font, fill = _get_font_n_fill(font_color, bg_color, bold)
    if font is not None and fill is not None:
        dxf = xl_diff_styles.DifferentialStyle(font=font, fill=fill)
    elif font is not None:
        dxf = xl_diff_styles.DifferentialStyle(font=font)
    elif fill is not None:
        dxf = xl_diff_styles.DifferentialStyle(fill=fill)
    else:
        dxf = None
    return dxf


def add_formula_conditional_formatting(ws: openpyxl.worksheet.worksheet.Worksheet, cell_range: str, formula: str,
                                       font_color: str, bg_color: str, bold: bool = False) -> None:
    font, fill = _get_font_n_fill(font_color, bg_color, bold)
    if font is not None and fill is not None:
        formula_rule = xl_format_rule.FormulaRule(formula=[formula], stopIfTrue=True, font=font, fill=fill)
    elif font is not None:
        formula_rule = xl_format_rule.FormulaRule(formula=[formula], stopIfTrue=True, font=font)
    elif fill is not None:
        formula_rule = xl_format_rule.FormulaRule(formula=[formula], stopIfTrue=True, fill=fill)
    else:
        return
    ws.conditional_formatting.add(cell_range, formula_rule)


# ------ Conditional Formatting ------ #

def create_defined_name(wb: openpyxl.workbook.Workbook, sheet_name: str, cell_range: str, name: str):
    ref = f"{xl_ut.quote_sheetname(sheet_name)}!{xl_ut.absolute_coordinate(cell_range)}"
    defined_name = xl_defined_name.DefinedName(name, attr_text=ref)
    wb.defined_names.add(defined_name)
