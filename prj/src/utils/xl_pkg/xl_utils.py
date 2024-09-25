#!/usr/bin/env python
# -*- coding: utf-8 -*-

import warnings
import openpyxl
import openpyxl.utils as xl_ut
import openpyxl.worksheet.worksheet as xl_ws
import xlrd
from ..common_utils import *


__all__ = ["openpyxl", "xlrd", "xl_ut", "xl_ws", "load_xlsx_wb", "load_xlrd_wb", "get_xlrd_column", "get_xlrd_row",
           "get_xl_column_letter", "get_xl_column_number", "get_xlrd_float_value", "get_xlrd_value", "get_xlsx_value",
           "get_row_and_column_from_cell", "get_cell_from_row_and_column", "get_cell_range"]


def load_xlsx_wb(xl_file_address: str, template: bool = False, read_only: bool = False) -> openpyxl.workbook.Workbook:
    warnings.filterwarnings("ignore", message="Cannot parse header or footer so it will be ignored",
                            category=UserWarning, module="openpyxl")  # deactivate the warning for header/footer
    warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed",
                            category=UserWarning, module="openpyxl")  # deactivate the warning for extension
    if template:
        wb = openpyxl.load_workbook(xl_file_address, rich_text=True)
    elif read_only:
        wb = openpyxl.load_workbook(xl_file_address, data_only=True, read_only=True, keep_links=False)
    else:
        # read-only active prevents the max_row to work
        wb = openpyxl.load_workbook(xl_file_address, data_only=True, read_only=False, keep_links=False)
    warnings.filterwarnings("default")
    return wb


def load_xlrd_wb(xl_file_address: str, formatting_info=False, on_demand=False) -> openpyxl.workbook.Workbook:
    # formatting_info is useful if a cell is in percentage mode
    wb = xlrd.open_workbook(xl_file_address, formatting_info=formatting_info, on_demand=on_demand)
    return wb


def get_xlrd_column(column_number: int) -> int:
    return column_number - 1


def get_xlrd_row(row: int) -> int:
    return row - 1


def get_xl_column_letter(column_number: int) -> str:
    return xl_ut.get_column_letter(column_number)


def get_xl_column_number(column_letter: str) -> int:
    return xl_ut.column_index_from_string(column_letter)


def get_xlrd_float_value(ws: xlrd.sheet, row: int, column: int) -> Optional[Union[float, str]]:
    value = get_xlrd_value(ws, row, column)
    if isinstance(value, str):
        try:
            value = float(value.replace(",", "."))
        except ValueError:
            pass
    return value


def get_xlrd_value(ws: xlrd.sheet, row: int, column: int) -> Optional[str]:
    xlrd_row = get_xlrd_row(row)
    xlrd_col = get_xlrd_column(column)
    try:
        cell_value = ws.cell_value(xlrd_row, xlrd_col)
    except IndexError:
        cell_value = None
    if cell_value == "":
        cell_value = None
    return cell_value


def get_xlsx_value(ws, row: int, column: int) -> Optional[str]:
    cell_value = ws.cell(row=row, column=column).value
    if cell_value == "":
        cell_value = None
    return cell_value


def get_row_and_column_from_cell(cell: str = None, row: int = None, column: Union[str, int] = None) -> tuple[int, int]:
    if cell is not None:
        row, column = split_cell(cell)
    if isinstance(column, str):
        column = xl_ut.column_index_from_string(column)
    return row, column


def split_cell(cell: str) -> tuple[int, str]:
    first_digit_index = -1
    for i, character in enumerate(cell):
        if character.isdigit():
            first_digit_index = i
            break
    column = cell[:first_digit_index]
    row = int(cell[first_digit_index:])
    return row, column


def get_cell_from_row_and_column(cell: str = None, row: int = None, column: Union[str, int] = None) -> str:
    if cell is None:
        if isinstance(column, int):
            column = xl_ut.get_column_letter(column)
        cell = f"{column}{row}"
    return cell


def get_cell_range(cell_range: str = None, start_row: int = None, end_row: int = None,
                   start_column: Union[int, str] = None, end_column: Union[int, str] = None) -> str:
    if cell_range is None:
        start_cell = get_cell_from_row_and_column(row=start_row, column=start_column)
        end_cell = get_cell_from_row_and_column(row=end_row, column=end_column)
        cell_range = f"{start_cell}:{end_cell}"
    return cell_range
