#!/usr/bin/env python
# -*- coding: utf-8 -*-

import typing
import openpyxl
import openpyxl.utils as xl_ut
import openpyxl.comments as xl_comments
import xlrd


def get_xlrd_column(col_nb: int) -> int:
    return col_nb - 1


def get_xlrd_line(line: int) -> int:
    return line - 1


def load_xlsx_wb(path: str) -> openpyxl.workbook.Workbook:
    return openpyxl.load_workbook(path)


def get_xl_column(col: int) -> str:
    return xl_ut.get_column_letter(col + 1)


def get_xl_line(line: int) -> int:
    return line + 1


def get_xlrd_float_value(sh: xlrd.sheet, line: int, col: int):
    value = get_xlrd_value(sh, line, col)
    if isinstance(value, str):
        try:
            value = float(value.replace(",", "."))
        except ValueError:
            pass
    return value


def get_xlrd_value(sh: xlrd.sheet, line: int, col: int) -> str:
    xlrd_line = get_xlrd_line(line)
    xlrd_col = get_xlrd_column(col)
    try:
        cell_value = sh.cell_value(xlrd_line, xlrd_col)
    except IndexError:
        cell_value = None
    if cell_value == "":
        cell_value = None
    return cell_value


def get_xlsx_value(sh, line: int, col: int) -> str:
    cell_value = sh.cell(row=line, column=col).value
    if cell_value == "":
        cell_value = None
    return cell_value


def get_cell_line_col(cell: str = None, line: int = None, col: typing.Union[str, int] = None):
    if cell is not None:
        line = int(cell[1])
        col = cell[0]
    if isinstance(col, str):
        col = xl_ut.column_index_from_string(col)
    return line, col


def get_line_col_cell(cell: str = None, line: int = None, col: typing.Union[str, int] = None):
    if cell is None:
        if isinstance(col, int):
            col = xl_ut.get_column_letter(col)
        cell = f"{col}{line}"
    return cell


def add_cell_comment(sh, comment: str, cell: str = None, line: int = None, col: typing.Union[str, int] = None):
    cell = get_line_col_cell(cell, line, col)
    comment = xl_comments.Comment(comment, None)
    sh[cell].comment = comment
